import streamlit as st
import streamlit.components.v1 as components
import requests, json, time, base64, re, hashlib
from html import escape as _esc
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from itens_checklist import ITENS_CHECKLIST, ITENS_VAN

# ─────────────────────────────────────────────
st.set_page_config(page_title="Checklist Veículos", page_icon="🚗",
                   layout="wide", initial_sidebar_state="collapsed")

TENANT_ID     = st.secrets["TENANT_ID"]
CLIENT_ID     = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]
SITE_HOST     = st.secrets.get("SITE_HOST", "metalcana.sharepoint.com")
SITE_NAME     = st.secrets.get("SITE_NAME", "AppKanbanFrotas")
LISTA_FROTAS      = st.secrets.get("LISTA_FROTAS",      "KanbanFrotas")
LISTA_CHECKLIST   = st.secrets.get("LISTA_CHECKLIST",   "ChecklistVeiculos")
LISTA_FOTOS       = st.secrets.get("LISTA_FOTOS",       "ChecklistFotos")
LISTA_PECAS       = st.secrets.get("LISTA_PECAS",       "SolicitacaoPecas")

TIPOS_VEICULO = ["Carro", "Caminhonete", "Van / Sprinter"]
AVALIACOES    = ["✅ Bom", "⚠️ Regular", "❌ Ruim", "➖ Não tem"]
AVAL_COR      = {"✅ Bom": "#d1fae5", "⚠️ Regular": "#fef3c7", "❌ Ruim": "#fee2e2", "➖ Não tem": "#f3f4f6"}
AVAL_TX       = {"✅ Bom": "#065f46", "⚠️ Regular": "#92400e", "❌ Ruim": "#991b1b", "➖ Não tem": "#4b5563"}

# ─────────────────────────────────────────────
# CODEC DO CHECKLIST
# O JSON com os 69 itens escritos por extenso dá ~4300 caracteres e a coluna
# "Itens" aceita 4000 — salvar cortado quebrava o JSON e o histórico não
# conseguia ler nenhum item. Aqui gravamos 1 caractere por item, na ordem de
# itens_checklist.py, o que dá ~950 caracteres e cabe com folga.
# ─────────────────────────────────────────────
AVAL_COD = {"✅ Bom": "B", "⚠️ Regular": "R", "❌ Ruim": "X", "➖ Não tem": "N"}
COD_AVAL = {v: k for k, v in AVAL_COD.items()}


def flat_itens(tipo_veiculo=""):
    """Lista achatada (categoria, item) na mesma ordem em que o checklist é exibido."""
    cats = dict(ITENS_CHECKLIST)
    if tipo_veiculo == "Van / Sprinter":
        cats.update(ITENS_VAN)
    return [(cat, item) for cat, lst in cats.items() for item in lst]


def hash_itens(flat):
    """Impressão digital da lista de itens — detecta se o template mudou
    depois que um checklist antigo foi salvo."""
    base = "|".join(f"{c}::{i}" for c, i in flat)
    return hashlib.md5(base.encode("utf-8")).hexdigest()[:8]


def codificar_itens(itens_state, tipo_veiculo=""):
    flat = flat_itens(tipo_veiculo)
    mask, obs = "", {}
    for idx, (cat, item) in enumerate(flat):
        mask += AVAL_COD.get(itens_state.get(f"{cat}||{item}", ""), "-")
        o = (itens_state.get(f"{cat}||{item}__obs", "") or "").strip()
        if o:
            obs[str(idx)] = o[:300]
    payload = {"v": 2, "h": hash_itens(flat), "tipo": tipo_veiculo, "m": mask, "o": obs}
    txt = json.dumps(payload, ensure_ascii=False)
    while len(txt) > 3900 and obs:      # se estourar, corta observação — nunca o gabarito
        obs.pop(max(obs, key=lambda k: len(obs[k])))
        payload["o"] = obs
        txt = json.dumps(payload, ensure_ascii=False)
    return txt


def _completar(avals_raw, obs_raw, tipo_veiculo):
    """Encaixa o que foi lido na lista completa de itens; o que faltar fica vazio."""
    flat = flat_itens(tipo_veiculo)
    avals = {f"{c}||{i}": avals_raw.get(f"{c}||{i}", "") for c, i in flat}
    for k, v in avals_raw.items():
        if k not in avals:
            avals[k] = v
    return avals, dict(obs_raw)


def _split_antigo(d, tipo_veiculo):
    avals = {k: v for k, v in d.items() if not k.endswith("__obs")}
    obs   = {k[:-5]: v for k, v in d.items() if k.endswith("__obs")}
    return _completar(avals, obs, tipo_veiculo)


def decodificar_itens(raw, tipo_veiculo=""):
    """Devolve (avaliações, observações, aviso). Entende o formato novo (v2),
    o JSON antigo completo e o JSON antigo cortado no meio."""
    if not raw:
        return {}, {}, ""
    try:
        d = json.loads(raw)
        if isinstance(d, dict) and d.get("v") == 2:
            tipo  = d.get("tipo", tipo_veiculo)
            flat  = flat_itens(tipo)
            aviso = ""
            if d.get("h") and d["h"] != hash_itens(flat):
                aviso = ("A lista de itens mudou depois que este checklist foi salvo — "
                         "as avaliações abaixo podem estar deslocadas.")
            mask    = d.get("m", "")
            obs_pos = d.get("o", {}) or {}
            avals, obs = {}, {}
            for idx, (cat, item) in enumerate(flat):
                chave = f"{cat}||{item}"
                avals[chave] = COD_AVAL.get(mask[idx] if idx < len(mask) else "-", "")
                if obs_pos.get(str(idx)):
                    obs[chave] = obs_pos[str(idx)]
            return avals, obs, aviso
        if isinstance(d, dict):
            a, o = _split_antigo(d, tipo_veiculo)
            return a, o, ""
    except Exception:
        pass
    # JSON cortado no limite da coluna — recupera os pares que couberam
    pares = dict(re.findall(r'"([^"]+?)"\s*:\s*"([^"]*)"', raw))
    a, o = _split_antigo(pares, tipo_veiculo)
    faltando = sum(1 for v in a.values() if not v)
    aviso = (f"Checklist salvo antes da correção do limite da coluna: {faltando} item(ns) "
             "não chegaram a ser gravados no SharePoint.") if faltando else ""
    return a, o, aviso

# ─────────────────────────────────────────────
# GRAPH API
# ─────────────────────────────────────────────
@st.cache_data(ttl=3500)
def get_token():
    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={"grant_type":"client_credentials","client_id":CLIENT_ID,
              "client_secret":CLIENT_SECRET,"scope":"https://graph.microsoft.com/.default"})
    r.raise_for_status()
    return r.json()["access_token"]

@st.cache_data(ttl=3500)
def get_site_id():
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{SITE_HOST}:/sites/{SITE_NAME}",
        headers={"Authorization": f"Bearer {get_token()}"})
    r.raise_for_status()
    return r.json()["id"]

def hdrs():
    return {"Authorization": f"Bearer {get_token()}", "Content-Type": "application/json"}

def _checar(r):
    """Levanta erro com o corpo da resposta da Graph API (mensagem real do problema)."""
    if not r.ok:
        try:
            detalhe = r.json().get("error", {}).get("message", r.text)
        except Exception:
            detalhe = r.text
        raise Exception(f"Graph API {r.status_code}: {detalhe}")

def lista_items(lista, filtro=""):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista}/items?expand=fields&$top=500"
    if filtro:
        url += f"&$filter={filtro}"
    itens = []
    while url:
        r = requests.get(url, headers=hdrs())
        _checar(r)
        data = r.json()
        itens.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return itens

def criar_item(lista, fields):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista}/items"
    r = requests.post(url, headers=hdrs(), json={"fields": fields})
    _checar(r)
    return r.json()

def patch_item(lista, item_id, fields):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista}/items/{item_id}/fields"
    r = requests.patch(url, headers=hdrs(), json=fields)
    _checar(r)
    return r.json()

def upload_foto_biblioteca(frota_nome, item_nome, foto_bytes):
    """Sobe a foto como arquivo de verdade na biblioteca de documentos do site
    (aparece com miniatura navegando direto no SharePoint) e devolve a URL."""
    site_id = get_site_id()
    pasta = re.sub(r'[\\/:*?"<>|]', "-", frota_nome).strip()
    nome_item = re.sub(r'[\\/:*?"<>|]', "-", item_nome).strip()[:60]
    nome_arquivo = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nome_item}.jpg"
    caminho = f"FotosChecklist/{pasta}/{nome_arquivo}"
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{caminho}:/content"
    r = requests.put(
        url,
        headers={"Authorization": f"Bearer {get_token()}", "Content-Type": "image/jpeg"},
        data=foto_bytes)
    _checar(r)
    j = r.json()
    return {"url": j.get("webUrl", ""), "id": j.get("id", "")}


def _share_id(url: str) -> str:
    """Converte a webUrl do arquivo no formato de shareId aceito pelo Graph."""
    return "u!" + base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")


@st.cache_data(ttl=1800, show_spinner=False)
def baixar_foto(foto_url: str) -> bytes:
    """Baixa os bytes reais da foto que está na biblioteca de documentos.
    A webUrl do SharePoint exige login interativo, então ela nao pode ser usada
    direto no st.image — aqui resolvemos o arquivo pelo Graph (token do app) e
    devolvemos a imagem completa."""
    if not foto_url:
        return b""
    try:
        r = requests.get(
            f"https://graph.microsoft.com/v1.0/shares/{_share_id(foto_url)}/driveItem",
            headers={"Authorization": f"Bearer {get_token()}"}, timeout=30)
        if r.status_code != 200:
            return b""
        download_url = r.json().get("@microsoft.graph.downloadUrl", "")
        if not download_url:
            return b""
        rc = requests.get(download_url, timeout=60)
        return rc.content if rc.status_code == 200 else b""
    except Exception:
        return b""


def _b64_legado(txt: str) -> bytes:
    """Fallback para registros antigos: o base64 foi cortado em 3900 chars,
    então recuperamos o pedaço válido (a imagem vem incompleta)."""
    if not txt:
        return b""
    try:
        return base64.b64decode(txt[: len(txt) // 4 * 4])
    except Exception:
        return b""

# ─────────────────────────────────────────────
# CARREGAR FROTAS DO KANBAN
# ─────────────────────────────────────────────
@st.cache_data(ttl=60)
def carregar_frotas():
    items = lista_items(LISTA_FROTAS)
    return [{
        "id":          i["id"],
        "nome":        i["fields"].get("Title",""),
        "tipo":        i["fields"].get("Tipo",""),
        "chassi":      i["fields"].get("Chassi",""),
        "ano":         i["fields"].get("Ano",""),
        "cc_nome":     i["fields"].get("CCNome",""),
        "frente_nome": i["fields"].get("FrenteNome",""),
        "status":      i["fields"].get("Status","Ativo"),
    } for i in items]

@st.cache_data(ttl=30)
def carregar_checklists(frota_id=""):
    items = lista_items(LISTA_CHECKLIST)
    cl = [{
        "id":            i["id"],
        "frota_nome":    i["fields"].get("FrotaNome",""),
        "frota_id":      i["fields"].get("FrotaId",""),
        "cc_nome":       i["fields"].get("CCNome",""),
        "frente_nome":   i["fields"].get("FrenteNome",""),
        "tipo_veiculo":  i["fields"].get("TipoVeiculo",""),
        "chassi":        i["fields"].get("Chassi",""),
        "ano":           i["fields"].get("Ano",""),
        "data":          i["fields"].get("DataChecklist",""),
        "operador":      i["fields"].get("NomeOperador",""),
        "inspetor":      i["fields"].get("NomeInspetor",""),
        "km":            i["fields"].get("KmAtual",""),
        "motivo":        i["fields"].get("MotivoEntrada",""),
        "resultado":     i["fields"].get("Resultado",""),
        "itens":         i["fields"].get("Itens","{}"),
        "obs":           i["fields"].get("Observacoes",""),
        "status":        i["fields"].get("Status","Rascunho"),
        "created":       i["fields"].get("Created",""),
    } for i in items]
    if frota_id:
        cl = [c for c in cl if c["frota_id"] == frota_id]
    return sorted(cl, key=lambda x: x.get("created",""), reverse=True)

def invalidar():
    carregar_frotas.clear()
    carregar_checklists.clear()
    carregar_pecas.clear()
    carregar_fotos.clear()

@st.cache_data(ttl=30)
def carregar_fotos(checklist_id=""):
    items = lista_items(LISTA_FOTOS)
    fotos = [{
        "id":           i["id"],
        "checklist_id": i["fields"].get("ChecklistId",""),
        "frota_nome":   i["fields"].get("FrotaNome",""),
        "item":         i["fields"].get("Item",""),
        "categoria":    i["fields"].get("Categoria",""),
        "avaliacao":    i["fields"].get("Avaliacao",""),
        "foto_base64":  i["fields"].get("FotoBase64",""),
        "foto_url":     i["fields"].get("FotoUrl",""),
        "observacao":   i["fields"].get("Observacao",""),
    } for i in items]
    if checklist_id:
        fotos = [f for f in fotos if f["checklist_id"] == checklist_id]
    return fotos

@st.cache_data(ttl=30)
def carregar_pecas(checklist_id=""):
    items = lista_items(LISTA_PECAS)
    pecas = [{
        "id":               i["id"],
        "checklist_id":     i["fields"].get("ChecklistId",""),
        "checklist_titulo": i["fields"].get("ChecklistTitulo",""),
        "frota_nome":       i["fields"].get("FrotaNome",""),
        "frota_id":         i["fields"].get("FrotaId",""),
        "cc_nome":          i["fields"].get("CCNome",""),
        "inspetor":         i["fields"].get("NomeInspetor",""),
        "nome_peca":        i["fields"].get("NomePeca",""),
        "quantidade":       i["fields"].get("Quantidade",1),
        "urgencia":         i["fields"].get("Urgencia","Normal"),
        "observacao":       i["fields"].get("Observacao",""),
        "status":           i["fields"].get("Status","Solicitado"),
        "comprador":        i["fields"].get("NomeComprador",""),
        "data_compra":      i["fields"].get("DataCompra",""),
        "data_recebimento": i["fields"].get("DataRecebimento",""),
        "data_instalacao":  i["fields"].get("DataInstalacao",""),
        "valor_unitario":   i["fields"].get("ValorUnitario",0),
        "obs_comprador":    i["fields"].get("ObsComprador",""),
        "created":          i["fields"].get("Created",""),
    } for i in items]
    if checklist_id:
        pecas = [p for p in pecas if p["checklist_id"] == checklist_id]
    return sorted(pecas, key=lambda x: x["created"], reverse=True)

STATUS_PECAS  = ["Solicitado","Comprado","Recebido","Instalado"]
URGENCIA_OPTS = ["Normal","Urgente","Crítico"]
URGENCIA_COR  = {"Normal":"#dbeafe","Urgente":"#fef3c7","Crítico":"#fee2e2"}
URGENCIA_TX   = {"Normal":"#1e40af","Urgente":"#92400e","Crítico":"#991b1b"}
STATUS_COR    = {
    "Solicitado": "#f3f4f6", "Comprado":  "#dbeafe",
    "Recebido":   "#d1fae5", "Instalado": "#d1fae5",
}
STATUS_TX     = {
    "Solicitado": "#374151", "Comprado":  "#1e40af",
    "Recebido":   "#065f46", "Instalado": "#065f46",
}
STATUS_EMOJI  = {
    "Solicitado":"🟡","Comprado":"🔵","Recebido":"🟢","Instalado":"✅"
}



# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
#MainMenu, footer, header {visibility:hidden;}
.block-container {padding:1rem 1.5rem !important;}
.item-row {
    display:flex; align-items:center; justify-content:space-between;
    padding:8px 12px; border-radius:8px; margin-bottom:6px;
    background:#fff; border:1px solid #e5e7eb;
}
.item-nome {font-size:13px; color:#111; font-weight:500;}
.cat-header {
    font-size:13px; font-weight:700; padding:8px 12px;
    border-radius:8px; margin:12px 0 6px; color:#fff;
}
.resultado-box {
    border-radius:10px; padding:16px 20px; text-align:center;
    font-size:20px; font-weight:700; margin:8px 0;
}
.hist-card {
    background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px;
    padding:12px 16px; margin-bottom:10px;
}
.gab-cat {
    font-size:12px; font-weight:700; color:#fff; background:#4b5563;
    padding:5px 10px; border-radius:6px; margin:12px 0 2px;
}
.gab-row {
    display:flex; align-items:flex-start; gap:10px;
    padding:5px 8px; font-size:12.5px;
    border-bottom:1px solid rgba(128,128,128,.18);
}
.gab-num  {min-width:24px; color:#9ca3af; font-variant-numeric:tabular-nums;}
.gab-item {flex:1; line-height:1.35;}
.gab-obs  {display:block; font-size:11px; color:#9ca3af; font-style:italic; margin-top:2px;}
.gab-badge{
    white-space:nowrap; border-radius:6px; padding:2px 8px;
    font-size:11px; font-weight:600;
}
.gab-chip {
    display:inline-block; border-radius:12px; padding:3px 10px;
    margin:0 6px 6px 0; font-size:11.5px; font-weight:600;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# ESTADO DA SESSÃO
# ─────────────────────────────────────────────
if "checklist_itens" not in st.session_state:
    st.session_state.checklist_itens = {}
if "fotos" not in st.session_state:
    st.session_state.fotos = {}
if "etapa" not in st.session_state:
    st.session_state.etapa = "identificacao"  # identificacao | checklist | revisao | concluido
if "frota_selecionada" not in st.session_state:
    st.session_state.frota_selecionada = None
if "tipo_veiculo" not in st.session_state:
    st.session_state.tipo_veiculo = "Carro"

# ─────────────────────────────────────────────
# TÍTULO E ABAS
# ─────────────────────────────────────────────
st.markdown("### 🚗 Checklist de Veículos — Teston / Metalcana")
aba_novo, aba_hist, aba_pecas, aba_export = st.tabs([
    "📋 Novo Checklist",
    "🕐 Histórico",
    "🔩 Peças",
    "📤 Exportar",
])

# ══════════════════════════════════════════════
# ABA 1 — NOVO CHECKLIST
# ══════════════════════════════════════════════
with aba_novo:

    # ── PROGRESSO ────────────────────────────
    etapas = {"identificacao": 1, "checklist": 2, "revisao": 3, "concluido": 4}
    pct = {1: 25, 2: 60, 3: 90, 4: 100}
    etapa_atual = st.session_state.etapa
    prog = pct.get(etapas.get(etapa_atual, 1), 25)
    labels = ["1. Identificação", "2. Checklist", "3. Revisão", "4. Concluído"]
    col_prog = st.columns(4)
    for i, (col, lbl) in enumerate(zip(col_prog, labels)):
        ativo = etapas.get(etapa_atual, 1) >= i+1
        col.markdown(
            f'<div style="text-align:center;font-size:12px;font-weight:{"700" if ativo else "400"};'
            f'color:{"#1D9E75" if ativo else "#9ca3af"}">{lbl}</div>',
            unsafe_allow_html=True)
    st.progress(prog / 100)
    st.markdown("<br>", unsafe_allow_html=True)

    # ════════════════════════════════════════
    # ETAPA 1 — IDENTIFICAÇÃO
    # ════════════════════════════════════════
    if etapa_atual == "identificacao":
        st.subheader("1️⃣ Identificação do veículo e responsáveis")

        frotas = carregar_frotas()
        nomes_frotas = sorted([f["nome"] for f in frotas])

        col1, col2 = st.columns([2, 1])
        with col1:
            busca_frota = st.text_input("🔍 Buscar frota pelo código ou nome",
                                         placeholder="Ex: 1185 ou TOYOTA")
        with col2:
            tipo_v = st.selectbox("Tipo de veículo", TIPOS_VEICULO)

        frotas_filtradas = [f for f in frotas
                            if not busca_frota or busca_frota.lower() in f["nome"].lower()]
        nomes_filtrados = ["— Selecione —"] + [f["nome"] for f in frotas_filtradas]

        frota_nome = st.selectbox("Frota", nomes_filtrados)
        frota_obj  = next((f for f in frotas if f["nome"] == frota_nome), None)

        if frota_obj:
            st.markdown("**📍 Dados do Kanban:**")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Centro de Custo", frota_obj["cc_nome"] or "—")
            c2.metric("Frente",          frota_obj["frente_nome"] or "—")
            c3.metric("Chassi",          frota_obj["chassi"] or "—")
            c4.metric("Ano",             frota_obj["ano"] or "—")

        st.divider()
        c1,c2,c3 = st.columns(3)
        with c1: operador = st.text_input("👤 Nome do operador / motorista")
        with c2: inspetor  = st.text_input("🔍 Nome do inspetor")
        with c3: km_atual  = st.number_input("🛣️ KM atual", min_value=0, step=1)

        motivo = st.selectbox("📝 Motivo da entrada na oficina", [
            "Manutenção corretiva",
            "Revisão periódica", "Checklist de rotina",
            "Após acidente / sinistro", "Outro",
        ])
        if motivo == "Outro":
            motivo = st.text_input("Descreva o motivo")

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("▶️ Iniciar Checklist", type="primary", use_container_width=True):
            if frota_nome == "— Selecione —":
                st.error("Selecione uma frota.")
            elif not operador.strip():
                st.error("Informe o nome do operador.")
            elif not inspetor.strip():
                st.error("Informe o nome do inspetor.")
            else:
                st.session_state.frota_selecionada = frota_obj
                st.session_state.tipo_veiculo = tipo_v
                st.session_state.dados_ident = {
                    "operador": operador.strip(),
                    "inspetor": inspetor.strip(),
                    "km": km_atual,
                    "motivo": motivo,
                }
                st.session_state.checklist_itens = {}
                st.session_state.fotos = {}
                st.session_state.etapa = "checklist"
                st.rerun()

    # ════════════════════════════════════════
    # ETAPA 2 — CHECKLIST
    # ════════════════════════════════════════
    elif etapa_atual == "checklist":
        frota = st.session_state.frota_selecionada
        ident = st.session_state.dados_ident
        tipo_v = st.session_state.tipo_veiculo

        st.markdown(
            f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;'
            f'padding:10px 16px;margin-bottom:16px">'
            f'<b>🚗 {frota["nome"]}</b> &nbsp;|&nbsp; {frota["cc_nome"] or "Sem CC"} '
            f'&nbsp;|&nbsp; {tipo_v} &nbsp;|&nbsp; Operador: {ident["operador"]} '
            f'&nbsp;|&nbsp; Inspetor: {ident["inspetor"]}</div>',
            unsafe_allow_html=True)

        # Monta categorias
        cats = dict(ITENS_CHECKLIST)
        if tipo_v == "Van / Sprinter":
            cats.update(ITENS_VAN)

        CORES_CAT = ["#1D9E75","#378ADD","#BA7517","#D85A30","#7F77DD",
                     "#D4537E","#639922","#0F6E56","#185FA5"]

        itens_state = st.session_state.checklist_itens
        fotos_state = st.session_state.fotos

        for ci, (cat, itens) in enumerate(cats.items()):
            cor = CORES_CAT[ci % len(CORES_CAT)]
            st.markdown(
                f'<div class="cat-header" style="background:{cor}">{cat}</div>',
                unsafe_allow_html=True)

            for item in itens:
                key_aval = f"aval_{cat}_{item}"
                key_obs  = f"obs_{cat}_{item}"
                key_foto = f"foto_{cat}_{item}"

                c1, c2 = st.columns([3, 2])
                with c1:
                    st.markdown(f'<div style="font-size:13px;padding:8px 0">{item}</div>',
                                unsafe_allow_html=True)
                with c2:
                    aval = st.radio(
                        label=item, options=AVALIACOES,
                        key=key_aval, horizontal=True,
                        label_visibility="collapsed",
                        index=0,
                    )

                itens_state[f"{cat}||{item}"] = aval

                if aval == "❌ Ruim":
                    obs_col, foto_col = st.columns([2,1])
                    with obs_col:
                        obs = st.text_input(f"Observação — {item[:30]}...",
                                            key=key_obs, placeholder="Descreva o problema")
                        itens_state[f"{cat}||{item}__obs"] = obs
                    with foto_col:
                        foto = st.camera_input(f"📷 Foto (opcional)", key=key_foto)
                        if foto:
                            fotos_state[f"{cat}||{item}"] = base64.b64encode(foto.getvalue()).decode()
                            st.success("Foto capturada ✅")

                st.markdown('<hr style="margin:4px 0;border-color:#f3f4f6">', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        obs_geral = st.text_area("📝 Observações gerais (opcional)",
                                  placeholder="Anotações adicionais sobre o veículo...")
        st.session_state.obs_geral = obs_geral

        col_back, col_next = st.columns(2)
        with col_back:
            if st.button("◀️ Voltar", use_container_width=True):
                st.session_state.etapa = "identificacao"; st.rerun()
        with col_next:
            if st.button("▶️ Revisar e Finalizar", type="primary", use_container_width=True):
                st.session_state.etapa = "revisao"; st.rerun()

    # ════════════════════════════════════════
    # ETAPA 3 — REVISÃO
    # ════════════════════════════════════════
    elif etapa_atual == "revisao":
        frota  = st.session_state.frota_selecionada
        ident  = st.session_state.dados_ident
        itens  = st.session_state.checklist_itens
        fotos  = st.session_state.fotos
        obs_g  = st.session_state.get("obs_geral","")

        st.subheader("3️⃣ Revisão do checklist")

        # Calcula resultado
        avals = {k:v for k,v in itens.items() if not k.endswith("__obs")}
        n_bom  = sum(1 for v in avals.values() if v == "✅ Bom")
        n_reg  = sum(1 for v in avals.values() if v == "⚠️ Regular")
        n_ruim = sum(1 for v in avals.values() if v == "❌ Ruim")
        n_na   = sum(1 for v in avals.values() if v == "➖ Não tem")
        total  = len(avals)

        if n_ruim == 0 and n_reg == 0:
            resultado = "✅ Aprovado"
            res_cor   = "#d1fae5"; res_tx = "#065f46"
        elif n_ruim == 0:
            resultado = "⚠️ Aprovado com pendências"
            res_cor   = "#fef3c7"; res_tx = "#92400e"
        else:
            resultado = "❌ Reprovado — requer atenção"
            res_cor   = "#fee2e2"; res_tx = "#991b1b"

        # Resumo
        st.markdown(
            f'<div class="resultado-box" style="background:{res_cor};color:{res_tx}">'
            f'{resultado}</div>', unsafe_allow_html=True)

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Total de itens", total)
        c2.metric("✅ Bom", n_bom)
        c3.metric("⚠️ Regular", n_reg)
        c4.metric("❌ Ruim", n_ruim)
        c5.metric("➖ Não tem", n_na)

        st.divider()

        # Dados identificação
        st.markdown("**📋 Identificação**")
        ic1,ic2,ic3,ic4 = st.columns(4)
        ic1.markdown(f"**Frota:** {frota['nome']}")
        ic2.markdown(f"**CC:** {frota['cc_nome'] or '—'}")
        ic3.markdown(f"**Operador:** {ident['operador']}")
        ic4.markdown(f"**Inspetor:** {ident['inspetor']}")

        st.divider()

        # Itens com problema
        itens_problema = {k:v for k,v in avals.items() if v not in ("✅ Bom", "➖ Não tem")}
        if itens_problema:
            st.markdown("**⚠️ Itens com problema:**")
            for chave, aval in itens_problema.items():
                cat, item = chave.split("||")
                obs = itens.get(f"{chave}__obs","")
                tem_foto = chave in fotos
                bg = AVAL_COR[aval]; tx = AVAL_TX[aval]
                st.markdown(
                    f'<div style="background:{bg};border-radius:8px;padding:8px 12px;'
                    f'margin-bottom:6px;color:{tx}">'
                    f'<b>{cat}</b> → {item} &nbsp; <b>{aval}</b>'
                    f'{(" &nbsp;|&nbsp; " + obs) if obs else ""}'
                    f'{" &nbsp; 📷" if tem_foto else ""}'
                    f'</div>', unsafe_allow_html=True)
        else:
            st.success("Todos os itens estão em bom estado!")

        if obs_g:
            st.divider()
            st.markdown(f"**Observações gerais:** {obs_g}")

        # ── SOLICITAÇÃO DE PEÇAS ─────────────
        st.divider()
        st.markdown("**🔩 Solicitar peças para este checklist**")
        st.caption("Adicione as peças necessárias antes de finalizar.")

        if "pecas_solicitadas" not in st.session_state:
            st.session_state.pecas_solicitadas = []

        with st.form("form_add_peca", clear_on_submit=True):
            pc1, pc2, pc3, pc4 = st.columns([3,1,1,1])
            with pc1: nome_peca  = st.text_input("Nome da peça", placeholder="Ex: Pastilha de freio dianteira")
            with pc2: qtd_peca   = st.number_input("Qtd", min_value=1, value=1, step=1)
            with pc3: urg_peca   = st.selectbox("Urgência", URGENCIA_OPTS)
            with pc4:
                st.markdown("<br>", unsafe_allow_html=True)
                add_peca = st.form_submit_button("➕ Adicionar", use_container_width=True)
            obs_peca = st.text_input("Observação da peça (opcional)", placeholder="Referência, modelo específico...")
            if add_peca and nome_peca.strip():
                st.session_state.pecas_solicitadas.append({
                    "nome": nome_peca.strip(),
                    "quantidade": qtd_peca,
                    "urgencia": urg_peca,
                    "obs": obs_peca.strip(),
                })
                st.rerun()

        if st.session_state.pecas_solicitadas:
            for pi, peca in enumerate(st.session_state.pecas_solicitadas):
                urg_bg = URGENCIA_COR.get(peca["urgencia"],"#f3f4f6")
                urg_tx = URGENCIA_TX.get(peca["urgencia"],"#374151")
                pc_col, rm_col = st.columns([5,1])
                with pc_col:
                    st.markdown(
                        f'<div style="background:{urg_bg};border-radius:8px;padding:7px 12px;'
                        f'margin-bottom:4px;color:{urg_tx};font-size:13px">'
                        f'<b>{peca["nome"]}</b> &nbsp;|&nbsp; Qtd: {peca["quantidade"]} '
                        f'&nbsp;|&nbsp; {peca["urgencia"]}'
                        f'{(" &nbsp;|&nbsp; " + peca["obs"]) if peca["obs"] else ""}'
                        f'</div>', unsafe_allow_html=True)
                with rm_col:
                    if st.button("🗑️", key=f"rm_peca_{pi}"):
                        st.session_state.pecas_solicitadas.pop(pi)
                        st.rerun()
        else:
            st.info("Nenhuma peça adicionada ainda.")

        st.divider()
        cb, cc = st.columns(2)
        with cb:
            if st.button("◀️ Voltar ao checklist", use_container_width=True):
                st.session_state.etapa = "checklist"; st.rerun()
        with cc:
            if st.button("✅ Confirmar e Salvar", type="primary", use_container_width=True):
                with st.spinner("Salvando no SharePoint..."):
                    try:
                        # Salva checklist principal
                        item_json = codificar_itens(
                            itens, st.session_state.tipo_veiculo)
                        novo = criar_item(LISTA_CHECKLIST, {
                            "Title":         f"{frota['nome']} — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                            "FrotaNome":     frota["nome"],
                            "FrotaId":       frota["id"],
                            "CCNome":        frota["cc_nome"],
                            "FrenteNome":    frota["frente_nome"],
                            "TipoVeiculo":   st.session_state.tipo_veiculo,
                            "Chassi":        frota["chassi"],
                            "Ano":           frota["ano"],
                            "DataChecklist": datetime.now().isoformat(),
                            "NomeOperador":  ident["operador"],
                            "NomeInspetor":  ident["inspetor"],
                            "KmAtual":       ident["km"],
                            "MotivoEntrada": ident["motivo"],
                            "Resultado":     resultado,
                            "Itens":         item_json,  # já vem compacto, cabe na coluna
                            "Observacoes":   obs_g,
                            "Status":        "Concluído",
                        })
                        cl_id = novo["id"]

                        # Salva fotos
                        for chave, foto_b64 in fotos.items():
                            cat, item = chave.split("||")
                            aval = itens.get(chave,"")
                            obs_item = itens.get(f"{chave}__obs","")
                            try:
                                up = upload_foto_biblioteca(
                                    frota["nome"], item, base64.b64decode(foto_b64))
                                foto_url = up["url"]
                            except Exception as e:
                                foto_url = ""
                                st.warning(f"⚠️ A foto de '{item}' não subiu para a biblioteca: {e}")
                            criar_item(LISTA_FOTOS, {
                                "Title":       f"{frota['nome']} — {item[:50]}",
                                "ChecklistId": cl_id,
                                "FrotaNome":   frota["nome"],
                                "Item":        item,
                                "Categoria":   cat,
                                "Avaliacao":   aval,
                                # A coluna aceita no máximo 4000 chars — o base64 de uma foto
                                # tem centenas de milhares. Guardar cortado gerava imagem
                                # quebrada no histórico; a foto real fica na biblioteca.
                                "FotoBase64":  "",
                                "FotoUrl":     foto_url,
                                "Observacao":  obs_item,
                            })
                            time.sleep(0.1)

                        # Salva peças solicitadas
                        pecas_para_salvar = st.session_state.get("pecas_solicitadas",[])
                        for peca in pecas_para_salvar:
                            criar_item(LISTA_PECAS, {
                                "Title":            f"{frota['nome']} — {peca['nome'][:50]}",
                                "ChecklistId":      cl_id,
                                "ChecklistTitulo":  f"{frota['nome']} — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                                "FrotaNome":        frota["nome"],
                                "FrotaId":          frota["id"],
                                "CCNome":           frota["cc_nome"],
                                "NomeInspetor":     ident["inspetor"],
                                "NomePeca":         peca["nome"],
                                "Quantidade":       peca["quantidade"],
                                "Urgencia":         peca["urgencia"],
                                "Observacao":       peca["obs"],
                                "Status":           "Solicitado",
                            })
                            time.sleep(0.1)
                        st.session_state.pecas_solicitadas = []

                        invalidar()
                        st.session_state.ultimo_checklist_id = cl_id
                        st.session_state.ultimo_resultado    = resultado
                        st.session_state.etapa = "concluido"
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao salvar: {e}")

    # ════════════════════════════════════════
    # ETAPA 4 — CONCLUÍDO
    # ════════════════════════════════════════
    elif etapa_atual == "concluido":
        frota   = st.session_state.frota_selecionada
        result  = st.session_state.get("ultimo_resultado","")
        res_cor = "#d1fae5" if "Aprovado" in result and "pendências" not in result else \
                  "#fef3c7" if "pendências" in result else "#fee2e2"
        res_tx  = "#065f46" if "Aprovado" in result and "pendências" not in result else \
                  "#92400e" if "pendências" in result else "#991b1b"

        st.markdown(
            f'<div class="resultado-box" style="background:{res_cor};color:{res_tx};font-size:24px">'
            f'🎉 Checklist salvo com sucesso!<br>'
            f'<span style="font-size:16px">{result}</span></div>',
            unsafe_allow_html=True)
        st.markdown(f"**Veículo:** {frota['nome']}  |  "
                    f"**Data:** {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        st.divider()
        if st.button("➕ Novo Checklist", type="primary", use_container_width=True):
            st.session_state.etapa = "identificacao"
            st.session_state.checklist_itens = {}
            st.session_state.fotos = {}
            st.session_state.frota_selecionada = None
            st.rerun()

# ══════════════════════════════════════════════
# ABA 2 — HISTÓRICO
# ══════════════════════════════════════════════
with aba_hist:
    st.subheader("🕐 Histórico de Checklists")

    hf1, hf2, hf3 = st.columns([2,2,1])
    with hf1:
        busca_h = st.text_input("🔍 Buscar frota", placeholder="Nome ou código...", key="bh")
    with hf2:
        filtro_res = st.selectbox("Resultado", ["Todos","✅ Aprovado","⚠️ Pendências","❌ Reprovado"], key="fr")
    with hf3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Atualizar", use_container_width=True):
            invalidar(); st.rerun()

    try:
        todos_cl = carregar_checklists()
    except Exception as e:
        st.error(f"Erro ao carregar histórico: {e}")
        todos_cl = []

    # Filtra
    if busca_h:
        todos_cl = [c for c in todos_cl if busca_h.lower() in c["frota_nome"].lower()]
    if filtro_res != "Todos":
        todos_cl = [c for c in todos_cl if filtro_res.split()[1] in c["resultado"]]

    st.caption(f"{len(todos_cl)} checklist(s) encontrado(s)")
    st.divider()

    for cl in todos_cl[:50]:
        res = cl["resultado"]
        res_cor = "#d1fae5" if "Aprovado" in res and "pendências" not in res else \
                  "#fef3c7" if "pendências" in res else "#fee2e2"
        res_tx  = "#065f46" if "Aprovado" in res and "pendências" not in res else \
                  "#92400e" if "pendências" in res else "#991b1b"

        with st.expander(
            f"🚗 {cl['frota_nome']} — {cl['data'][:10] if cl['data'] else cl['created'][:10]} "
            f"| {res}"
        ):
            cc1,cc2,cc3,cc4 = st.columns(4)
            cc1.markdown(f"**CC:** {cl['cc_nome'] or '—'}")
            cc2.markdown(f"**Frente:** {cl['frente_nome'] or '—'}")
            cc3.markdown(f"**Operador:** {cl['operador']}")
            cc4.markdown(f"**Inspetor:** {cl['inspetor']}")

            st.markdown(f"**KM:** {cl['km']}  |  **Motivo:** {cl['motivo']}")

            # ── Gabarito completo do checklist ──────────
            try:
                fotos_cl = carregar_fotos(cl["id"])
            except Exception:
                fotos_cl = []
            com_foto = {f'{f["categoria"]}||{f["item"]}' for f in fotos_cl
                        if f["foto_url"] or f["foto_base64"]}

            avals, obs_itens, aviso = decodificar_itens(cl["itens"], cl["tipo_veiculo"])

            if avals:
                cont = {a: 0 for a in AVALIACOES}
                nao_reg = 0
                for v in avals.values():
                    if v in cont:
                        cont[v] += 1
                    else:
                        nao_reg += 1

                chips = "".join(
                    f'<span class="gab-chip" style="background:{AVAL_COR[a]};'
                    f'color:{AVAL_TX[a]}">{a}: {cont[a]}</span>' for a in AVALIACOES)
                if nao_reg:
                    chips += (f'<span class="gab-chip" style="background:#e5e7eb;'
                              f'color:#374151">Não registrado: {nao_reg}</span>')
                st.markdown(f'<div style="margin:8px 0 2px">{chips}</div>',
                            unsafe_allow_html=True)

                if aviso:
                    st.caption(f"⚠️ {aviso}")

                so_problemas = st.checkbox(
                    "Mostrar só os itens com problema",
                    key=f"gabfiltro_{cl['id']}")

                linhas, cat_atual, idx = [], None, 0
                for chave, aval in avals.items():
                    cat, _, item = chave.partition("||")
                    idx += 1
                    if so_problemas and aval not in ("⚠️ Regular", "❌ Ruim"):
                        continue
                    if cat != cat_atual:
                        cat_atual = cat
                        linhas.append(f'<div class="gab-cat">{_esc(cat)}</div>')
                    bg  = AVAL_COR.get(aval, "#e5e7eb")
                    tx  = AVAL_TX.get(aval, "#6b7280")
                    rot = aval if aval else "— não registrado"
                    cam = " 📷" if chave in com_foto else ""
                    o   = obs_itens.get(chave, "")
                    obs_html = f'<span class="gab-obs">{_esc(o)}</span>' if o else ""
                    linhas.append(
                        f'<div class="gab-row">'
                        f'<span class="gab-num">{idx:02d}</span>'
                        f'<span class="gab-item">{_esc(item)}{cam}{obs_html}</span>'
                        f'<span class="gab-badge" style="background:{bg};color:{tx}">{rot}</span>'
                        f'</div>')

                if linhas:
                    st.markdown("".join(linhas), unsafe_allow_html=True)
                else:
                    st.success("Nenhum item com problema neste checklist.")
            else:
                st.info("Este checklist não tem itens gravados.")

            if cl["obs"]:
                st.markdown(f"**Obs gerais:** {cl['obs']}")

            # Fotos deste checklist
            fotos_cl = [f for f in fotos_cl if f["foto_url"] or f["foto_base64"]]
            if fotos_cl:
                st.markdown("**📷 Fotos:**")
                cols_f = st.columns(min(len(fotos_cl), 4))
                for i, foto in enumerate(fotos_cl):
                    with cols_f[i % len(cols_f)]:
                        img = baixar_foto(foto["foto_url"]) or _b64_legado(foto["foto_base64"])
                        if img:
                            st.image(img, caption=foto["item"][:40], use_container_width=True)
                        else:
                            st.caption(f"📷 {foto['item'][:40]} — não foi possível carregar")
                        if foto["foto_url"]:
                            st.markdown(f"[Abrir no SharePoint]({foto['foto_url']})")

# ══════════════════════════════════════════════
# ABA 3 — PEÇAS
# ══════════════════════════════════════════════
with aba_pecas:
    st.subheader("🔩 Solicitação de Peças")

    # ── Filtros ──────────────────────────────
    pp1, pp2, pp3 = st.columns([2,2,1])
    with pp1:
        busca_p = st.text_input("🔍 Buscar frota ou peça", key="bp")
    with pp2:
        filtro_status_p = st.selectbox("Status", ["Todos"] + STATUS_PECAS, key="fsp")
    with pp3:
        filtro_urg_p = st.selectbox("Urgência", ["Todas"] + URGENCIA_OPTS, key="fup")

    pp4, pp5 = st.columns([2,1])
    with pp4:
        filtro_cc_p = st.selectbox("Centro de Custo", ["Todos os CCs"] + 
                                    list(dict.fromkeys([p["cc_nome"] for p in carregar_pecas() if p["cc_nome"]])), 
                                    key="fcp")
    with pp5:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Atualizar", key="ref_p", use_container_width=True):
            invalidar(); st.rerun()

    try:
        todas_pecas = carregar_pecas()
    except Exception as e:
        st.error(f"Erro: {e}"); todas_pecas = []

    if busca_p:
        todas_pecas = [p for p in todas_pecas 
                       if busca_p.lower() in p["frota_nome"].lower() 
                       or busca_p.lower() in p["nome_peca"].lower()]
    if filtro_status_p != "Todos":
        todas_pecas = [p for p in todas_pecas if p["status"] == filtro_status_p]
    if filtro_urg_p != "Todas":
        todas_pecas = [p for p in todas_pecas if p["urgencia"] == filtro_urg_p]
    if filtro_cc_p != "Todos os CCs":
        todas_pecas = [p for p in todas_pecas if p["cc_nome"] == filtro_cc_p]

    # Métricas rápidas
    m1,m2,m3,m4 = st.columns(4)
    for col, status, emoji in [
        (m1,"Solicitado","🟡"), (m2,"Comprado","🔵"),
        (m3,"Recebido","🟢"),   (m4,"Instalado","✅")
    ]:
        n = sum(1 for p in todas_pecas if p["status"] == status)
        col.markdown(
            f'<div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;'
            f'padding:8px 14px;text-align:center">'
            f'<div style="font-size:22px;font-weight:700">{n}</div>'
            f'<div style="font-size:11px;color:#6b7280">{emoji} {status}</div></div>',
            unsafe_allow_html=True)

    st.markdown(f"<br>*{len(todas_pecas)} peça(s) encontrada(s)*", unsafe_allow_html=True)
    st.divider()

    # ── Lista de peças ────────────────────────
    for peca in todas_pecas:
        urg_bg = URGENCIA_COR.get(peca["urgencia"],"#f3f4f6")
        urg_tx = URGENCIA_TX.get(peca["urgencia"],"#374151")
        st_bg  = STATUS_COR.get(peca["status"],"#f3f4f6")
        st_tx  = STATUS_TX.get(peca["status"],"#374151")
        emoji  = STATUS_EMOJI.get(peca["status"],"🟡")

        with st.expander(
            f'{emoji} **{peca["nome_peca"]}** — {peca["frota_nome"]} '
            f'| Qtd: {peca["quantidade"]} | {peca["urgencia"]} | {peca["status"]}'
        ):
            ec1,ec2,ec3 = st.columns(3)
            ec1.markdown(f"**Frota:** {peca['frota_nome']}")
            ec1.markdown(f"**CC:** {peca['cc_nome'] or '—'}")
            ec2.markdown(f"**Inspetor:** {peca['inspetor']}")
            ec2.markdown(f"**Checklist:** {peca['checklist_titulo']}")
            ec3.markdown(f"**Solicitado em:** {peca['created'][:10]}")
            if peca["observacao"]:
                ec3.markdown(f"**Obs:** {peca['observacao']}")

            st.markdown("**Atualizar status:**")
            with st.form(f"form_status_{peca['id']}"):
                sc1,sc2,sc3,sc4,sc5 = st.columns([2,1,1,2,2])
                with sc1:
                    novo_status = st.selectbox("Status", STATUS_PECAS,
                                               index=STATUS_PECAS.index(peca["status"]) 
                                               if peca["status"] in STATUS_PECAS else 0,
                                               key=f"ns_{peca['id']}")
                with sc2:
                    val_unit = st.number_input("Valor unit. R$", min_value=0.0,
                                               value=float(peca["valor_unitario"] or 0),
                                               step=0.01, key=f"vu_{peca['id']}")
                with sc3:
                    nome_comp = st.text_input("Comprador", value=peca["comprador"],
                                              key=f"nc_{peca['id']}")
                with sc4:
                    obs_comp = st.text_input("Obs do comprador", value=peca["obs_comprador"],
                                             key=f"oc_{peca['id']}")
                with sc5:
                    st.markdown("<br>", unsafe_allow_html=True)
                    salvar_st = st.form_submit_button("💾 Salvar", use_container_width=True)

                if salvar_st:
                    fields_upd = {
                        "Status":          novo_status,
                        "NomeComprador":   nome_comp,
                        "ValorUnitario":   val_unit,
                        "ObsComprador":    obs_comp,
                    }
                    # Preenche datas automáticas por status
                    hoje_iso = datetime.now().strftime("%Y-%m-%d")
                    if novo_status == "Comprado"  and not peca["data_compra"]:
                        fields_upd["DataCompra"]      = hoje_iso
                    if novo_status == "Recebido"  and not peca["data_recebimento"]:
                        fields_upd["DataRecebimento"] = hoje_iso
                    if novo_status == "Instalado" and not peca["data_instalacao"]:
                        fields_upd["DataInstalacao"]  = hoje_iso
                    try:
                        patch_item(LISTA_PECAS, peca["id"], fields_upd)
                        invalidar()
                        st.success("Status atualizado!"); st.rerun()
                    except Exception as e:
                        st.error(f"Erro: {e}")

# ══════════════════════════════════════════════
# ABA 4 — EXPORTAR
# ══════════════════════════════════════════════
with aba_export:
    st.subheader("📤 Exportar relatório de checklists")

    ex1, ex2 = st.columns([2,2])
    with ex1:
        busca_ex = st.text_input("Filtrar por frota", placeholder="Nome ou código...")
    with ex2:
        periodo  = st.selectbox("Período", ["Todos","Hoje","Últimos 7 dias","Últimos 30 dias"])

    try:
        todos_ex = carregar_checklists()
    except:
        todos_ex = []

    if busca_ex:
        todos_ex = [c for c in todos_ex if busca_ex.lower() in c["frota_nome"].lower()]
    if periodo == "Hoje":
        hoje = datetime.now().strftime("%Y-%m-%d")
        todos_ex = [c for c in todos_ex if c["created"][:10] == hoje]
    elif periodo == "Últimos 7 dias":
        from datetime import timedelta
        limite = (datetime.now() - timedelta(days=7)).isoformat()
        todos_ex = [c for c in todos_ex if c["created"] >= limite]
    elif periodo == "Últimos 30 dias":
        from datetime import timedelta
        limite = (datetime.now() - timedelta(days=30)).isoformat()
        todos_ex = [c for c in todos_ex if c["created"] >= limite]

    st.caption(f"{len(todos_ex)} checklist(s) para exportar")

    if st.button("⬇️ Gerar Excel", type="primary", use_container_width=True) and todos_ex:
        wb = Workbook(); wb.remove(wb.active)

        # Aba resumo
        ws = wb.create_sheet("Resumo")
        ws.merge_cells("A1:J1")
        ws["A1"] = f"RELATÓRIO DE CHECKLISTS — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True,size=12,color="FFFFFF")
        ws["A1"].fill = PatternFill("solid",fgColor="1D9E75")
        ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height = 26

        hdrs_r = ["Data","Frota","Tipo","CC","Frente","Operador","Inspetor","KM","Motivo","Resultado"]
        for ci,h in enumerate(hdrs_r,1): ws.cell(row=2,column=ci,value=h)
        fill_h = PatternFill("solid",fgColor="0F6E56")
        for ci in range(1,11):
            ws.cell(row=2,column=ci).fill  = fill_h
            ws.cell(row=2,column=ci).font  = Font(bold=True,color="FFFFFF",size=10)
            ws.cell(row=2,column=ci).alignment = Alignment(horizontal="center")

        for ri, cl in enumerate(todos_ex, 3):
            data_str = cl["data"][:10] if cl["data"] else cl["created"][:10]
            linha = [data_str, cl["frota_nome"], cl["tipo_veiculo"],
                     cl["cc_nome"], cl["frente_nome"], cl["operador"],
                     cl["inspetor"], cl["km"], cl["motivo"], cl["resultado"]]
            for ci,v in enumerate(linha,1): ws.cell(row=ri,column=ci,value=v)
            # Cor pelo resultado
            res_bg = "d1fae5" if "Aprovado" in cl["resultado"] and "pendências" not in cl["resultado"] else \
                     "fef3c7" if "pendências" in cl["resultado"] else "fee2e2"
            for ci in range(1,11):
                ws.cell(row=ri,column=ci).fill = PatternFill("solid",fgColor=res_bg)
                ws.cell(row=ri,column=ci).font = Font(size=10)
                ws.cell(row=ri,column=ci).border = Border(
                    left=Side("thin","E5E7EB"),right=Side("thin","E5E7EB"),
                    top=Side("thin","E5E7EB"),bottom=Side("thin","E5E7EB"))

        for ci,w in zip(range(1,11),[12,30,14,26,20,18,18,10,22,28]):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A3"

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        nome = f"checklists_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button("📥 Baixar Excel", buf, nome,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")